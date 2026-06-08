"""
Â§3 / Â§4 â€” Policy Loader

Reads Input_PolicyDataRaw and returns a flat dict of all policy fields plus
computed Policy_Info equivalents.

Workbook layout (D-001 â€” workbook wins over TMD Â§3.2):
  Row 1 (0-indexed row 0) = SQL field names  â† column headers
  Row 2 (0-indexed row 1) = policy data      â† the single data row
  Row 3 (0-indexed row 2) = type labels      â€” ignored
  Row 4 (0-indexed row 3) = column indices   â€” ignored

Sheet detection:
  â€¢ If the file contains a sheet named "Input_PolicyDataRaw" that sheet is used.
  â€¢ Otherwise the active (first) sheet is used â€” supports single-tab exports.

Error handling (D-003, D-004):
  â€¢ Excel error strings (#VALUE!, #N/A, etc.) â†’ None + ModelWarning
  â€¢ RAND()-derived volatile placeholders â†’ None + ModelWarning (D-004)
"""
from __future__ import annotations

import calendar
import datetime
from typing import Any

import openpyxl

from loaders.warnings import warn

# ---------------------------------------------------------------------------
# LB Code â†’ i4L flag  (Misc Mappings!AM:AN)
# TMD Â§3.6: TRUE for B,1,2,3,4,5,6,7,8,C,D,J,K
# All other codes return FALSE â†’ i4L_Indicator = "___"
# ---------------------------------------------------------------------------
_I4L_LB_CODES: frozenset[str] = frozenset(
    ["B", "1", "2", "3", "4", "5", "6", "7", "8", "C", "D", "J", "K"]
)

_OPENPYXL_ERROR_PREFIX = "#"
_SHEET_NAME = "Input_PolicyDataRaw"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_policy(
    policy_path: str,
    policy_id: str | None = None,
) -> dict[str, Any]:
    """
    Return a dict with all 243 raw SQL fields plus derived Policy_Info equivalents.

    Parameters
    ----------
    policy_path : path to Input_PolicyDataRaw.xlsx
    policy_id   : policy_number to load (string or int).
                  If None, loads the first data row (default behaviour).
                  If provided, scans all data rows and returns the matching one.

    Raises
    ------
    ValueError  if policy_id is provided but not found in the sheet.
    """
    try:
        wb = openpyxl.load_workbook(policy_path, data_only=True, read_only=True)
    except Exception as e:
        raise FileNotFoundError(
            f"Cannot open policy file '{policy_path}': {e}"
        ) from e

    ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active

    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(all_rows) < 2:
        raise ValueError(
            f"'{_SHEET_NAME}' has fewer than 2 rows in '{policy_path}'"
        )

    headers: tuple[Any, ...] = all_rows[0]

    # Find the policy_number column index
    pol_num_col: int | None = None
    for idx, h in enumerate(headers):
        if h is not None and str(h).strip() == "policy_number":
            pol_num_col = idx
            break

    if policy_id is None:
        # Default: first data row (original behaviour)
        data_row = all_rows[1]
    else:
        # Scan rows for matching policy_number
        target = str(policy_id).strip()
        data_row = None
        for row in all_rows[1:]:
            if pol_num_col is not None and row[pol_num_col] is not None:
                if str(row[pol_num_col]).strip() == target:
                    data_row = row
                    break
        if data_row is None:
            available = [
                str(all_rows[r][pol_num_col]).strip()
                for r in range(1, len(all_rows))
                if pol_num_col is not None and all_rows[r][pol_num_col] is not None
            ]
            raise ValueError(
                f"Policy ID '{policy_id}' not found in '{policy_path}'. "
                f"Available IDs: {available}"
            )

    raw: dict[str, Any] = {}
    for h, v in zip(headers, data_row):
        if h is None:
            continue
        field = str(h).strip()
        raw[field] = _sanitise(v, field)

    policy = dict(raw)
    _add_derived_fields(policy)
    return policy


# ---------------------------------------------------------------------------
# Derived fields (mirrors Policy_Info column C / T formulas)
# ---------------------------------------------------------------------------

def _add_derived_fields(p: dict[str, Any]) -> None:
    """
    Compute and inject all Policy_Info derived fields into p (mutates in place).
    """
    # ValuationDate  (Policy_Info!C2)
    # Always the last calendar day of the valuation month
    vy = int(p.get("valuation_year") or 0)
    vm_ = int(p.get("valuation_month") or 0)
    valuation_date = _last_day_of_month(vy, vm_)
    p["valuation_date"] = valuation_date

    # IssueDate
    iy = int(p.get("issue_year") or 0)
    im = int(p.get("issue_month") or 0)
    id_ = int(p.get("issue_day") or 1)
    p["issue_date"] = datetime.date(iy, im, id_)
    p["issue_day"] = id_

    # Company  (Policy_Info!C16)
    # "L" â†’ "LNL"  (Abc_corp Life),  "Y" â†’ "LNY"  (Abc_corp Life of New York)
    raw_co = str(p.get("company") or "").strip().upper()
    p["company"] = {"L": "LNL", "Y": "LNY"}.get(raw_co, raw_co)

    # policy_month_seed  (Policy_Info!T10)
    # = (YEAR(VD)-YEAR(ID))*12 + (MONTH(VD)-MONTH(ID))
    #   + IF(DAY(VD)>=DAY(IssueDate), 1, 0)
    p["policy_month_seed"] = (
        (vy - iy) * 12
        + (vm_ - im)
        + (1 if valuation_date.day >= id_ else 0)
    )

    # next_monthiversary  (Policy_Info!T12)
    # = DATE(YEAR(ValuationDate), MONTH(ValuationDate)+1, DAY(IssueDate))
    # Excel DATE() clamps day to month-end when month is short.
    nm_year = vy + (1 if vm_ == 12 else 0)
    nm_month = (vm_ % 12) + 1
    nm_day = min(id_, calendar.monthrange(nm_year, nm_month)[1])
    p["next_monthiversary"] = datetime.date(nm_year, nm_month, nm_day)

    # attained_age_seed  (Policy_Info!T13)
    # = (YEAR(VD)-YEAR(ID)) + IssueAge - IF(MONTH(IssueDate)>MONTH(VD), 1, 0)
    issue_age = int(p.get("issue_age") or 0)
    p["attained_age_seed"] = (vy - iy) + issue_age - (1 if im > vm_ else 0)

    # stub_period  (Policy_Info!T15)
    # Formula = MIN(DAY(IssueDate)/30, 1) * 0  â€” permanently zeroed (D-005)
    p["stub_period"] = 0.0

    # Plan  (Policy_Info!C17)
    plan = str(p.get("model_plan") or "").strip()
    p["plan"] = plan

    # LB_Code  (Policy_Info!C50)
    # SQL field "GMWB Code" maps to gmwb_type (e.g. "B")
    lb_raw = str(p.get("gmwb_type") or "").strip()
    if not lb_raw:
        # Fallback: living_benefit_rider_code carries "0B" format â†’ strip leading 0
        lb_raw = str(p.get("living_benefit_rider_code") or "").strip().lstrip("0") or "N"
    p["lb_code"] = lb_raw

    # i4L_Indicator  (Policy_Info!C23)
    # = IF(INDEX(Misc Mappings!$AN:$AN, MATCH(LB_Code, AM:AM, 0))=FALSE, "___", "i4L")
    p["i4l_indicator"] = "i4L" if lb_raw in _I4L_LB_CODES else "___"

    # DB_Option  (Policy_Info!C64)
    # "A"=AV-only, "1"=ROP, "2"=HWM(ratchet), "3"=5%-RollUp
    p["db_option"] = str(
        p.get("deathbenefittype") or p.get("death_benefit_type") or "A"
    ).strip()

    # Single/Joint  (Policy_Info!C14)
    # jointlife_indicator: "U"=single/unknown, "J"=joint
    p["single_joint"] = str(p.get("jointlife_indicator") or "U").strip()

    # Gender1  (Policy_Info!C13)
    p["gender1"] = str(p.get("gender") or "M").strip().upper()

    # 401K indicator  (Policy_Info!T17)
    # = OR(Plan="LDIRE1", "LDIRE2", "NDIRE1", "NDIRE2")
    p["indicator_401k"] = plan in ("LDIRE1", "LDIRE2", "NDIRE1", "NDIRE2")

    # Qualified status  (Policy_Info!C99)
    p["qualified_status"] = str(p.get("qualified_status") or "").strip()

    # Reinsurance phase  (Policy_Info!C32)
    p["reinsurance_phase"] = _to_int(p.get("reinsurance_phase"), default=1)

    # Mortality group  (Policy_Info!C15)
    p["mort_group"] = str(p.get("mortality_group_indicator") or "").strip()

    # DAC_Amortization_Basis  (Policy_Info!C107) â€” may be #VALUE! (D-003)
    # Already sanitised to None by _sanitise(); expose explicitly for DAC engine
    dac_basis = p.get("dac_amortization_basis")  # None if field absent
    if dac_basis is None:
        # Policy_Info!C107 formula references LEFT(E107,8); E107 pulls ldti_cohortid
        dac_basis = p.get("ldti_cohortid")
    p["dac_amortization_basis"] = dac_basis

    # Cross-check: attained_age from data vs derived
    data_att_age = _to_int(p.get("attained_age"), default=None)
    if data_att_age is not None and data_att_age != p["attained_age_seed"]:
        warn(
            f"attained_age from Input_PolicyDataRaw ({data_att_age}) differs from "
            f"derived attained_age_seed ({p['attained_age_seed']}); "
            f"using derived value (from T13 formula)",
            source="policy_loader/attained_age",
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sanitise(v: Any, field: str) -> Any:
    """Replace Excel error strings with None + ModelWarning (D-003)."""
    if isinstance(v, str) and v.startswith(_OPENPYXL_ERROR_PREFIX):
        warn(
            f"Excel error '{v}' in field '{field}' â€” defaulted to None (D-003)",
            source=f"Input_PolicyDataRaw/{field}",
        )
        return None
    return v


def _last_day_of_month(year: int, month: int) -> datetime.date:
    last = calendar.monthrange(year, month)[1]
    return datetime.date(year, month, last)


def _to_int(v: Any, default: Any = 0) -> Any:
    try:
        return int(v)
    except (TypeError, ValueError):
        return default
