"""
§6 — Scenario Loader

Reads ESG / interest-rate scenario data from two staging sheets inside the
policy workbook (Input_PolicyDataRaw.xlsx):

  • Input_Future Scenario   — 25 rows × 604 cols (600 projection months)
  • Input_Historic Scenario — 19 rows × 124 cols (120 lookback months)

Sheet layout (both sheets identical in structure):
  Row 1  : column headers ['No', 'Run Id', 'ScnName', 'Row', 'M1', 'M2', …]
            (Historic sheet uses '−M1', '−M2', … for month labels)
  Row 2+ : one variable per row
            Col 0 = No (int)
            Col 1 = Run Id (int)
            Col 2 = ScnName (str)
            Col 3 = Row / variable name (str)
            Col 4+ = monthly values

Returned ScenarioData:
  • future   : DataFrame(index=variable_name, columns=1..n_future_months)
  • historic : DataFrame(index=variable_name, columns=−1..−n_hist_months)
  • run_id, scn_name, no : metadata scalars from the first data row

Rate storage convention (source workbook):
  • Interest rates  (Interest-YC_*)  : percent BEY  (3.98 = 3.98%)
  • Inflation       (Inflation-*)    : percent       (3.98 = 3.98%)
  • Equity growth   (*-EQ_Growth)    : percent AEY   (11.58 = 11.58%)
  • Equity income   (*-EQ_Income)    : percent AEY
  • MRB user rate   (MRB_OCS-*)     : decimal        (0.0025 = 0.25%)
  Callers are responsible for unit conversion as needed.

Note: Paste_StdScn1, Paste_StdScn2, Paste_CARVM_PW, Paste_CARVM_NoPW are
output-staging areas (empty in the masked workbook). They are not loaded.
"""
from __future__ import annotations

import dataclasses
from typing import Any

import openpyxl
import pandas as pd

from loaders.warnings import warn

_FUTURE_SHEET  = "Input_Future Scenario"
_HISTORIC_SHEET = "Input_Historic Scenario"

# Column offsets in both sheets
_COL_NO      = 0
_COL_RUN_ID  = 1
_COL_SCN     = 2
_COL_ROW     = 3
_COL_FIRST_M = 4   # M1 / -M1 starts here


# ---------------------------------------------------------------------------
# Public data container
# ---------------------------------------------------------------------------

@dataclasses.dataclass(frozen=True)
class ScenarioData:
    """
    Holds parsed scenario data for one ESG run.

    future   : DataFrame(index=variable_name, columns=1..N)
               Monthly values for projection months 1..N (N ≤ 600).

    historic : DataFrame(index=variable_name, columns=−1..−M)
               Monthly values for lookback months −1..−M (M ≤ 120).
               Column −1 = one month before valuation date.

    run_id, scn_name, no : metadata read from the first data row.
    """
    run_id:   Any
    scn_name: Any
    no:       Any
    future:   pd.DataFrame
    historic: pd.DataFrame

    def get(self, variable: str) -> pd.Series:
        """Return the future-months Series for *variable*; raises KeyError if absent."""
        return self.future.loc[variable]

    def get_historic(self, variable: str) -> pd.Series:
        """Return the historic-months Series for *variable*; raises KeyError if absent."""
        return self.historic.loc[variable]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_scenarios(policy_path: str) -> ScenarioData:
    """
    Load scenario data from *policy_path* (full workbook or single-policy export).

    Returns a ScenarioData with:
      • future   — DataFrame(index=variable, columns=1..600)
      • historic — DataFrame(index=variable, columns=−1..−120)
      • run_id, scn_name, no scalars
    """
    try:
        wb = openpyxl.load_workbook(policy_path, data_only=True, read_only=True)
    except Exception as e:
        raise FileNotFoundError(
            f"Cannot open policy/scenario file '{policy_path}': {e}"
        ) from e

    future_df, run_id, scn_name, no = _parse_scenario_sheet(
        wb, _FUTURE_SHEET, historic=False
    )
    historic_df, *_ = _parse_scenario_sheet(
        wb, _HISTORIC_SHEET, historic=True
    )

    wb.close()

    return ScenarioData(
        run_id=run_id,
        scn_name=scn_name,
        no=no,
        future=future_df,
        historic=historic_df,
    )


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def _parse_scenario_sheet(
    wb,
    sheet_name: str,
    *,
    historic: bool,
) -> tuple[pd.DataFrame, Any, Any, Any]:
    """
    Parse one scenario sheet → (DataFrame, run_id, scn_name, no).

    DataFrame layout:
      index   = variable name (Row column)
      columns = month integers  (1..N  or  −1..−M)
    """
    if sheet_name not in wb.sheetnames:
        warn(
            f"Scenario sheet '{sheet_name}' not found in workbook — "
            f"returning empty DataFrame",
            source=f"scenario_loader/{sheet_name}",
        )
        return pd.DataFrame(), None, None, None

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        warn(
            f"Scenario sheet '{sheet_name}' is empty",
            source=f"scenario_loader/{sheet_name}",
        )
        return pd.DataFrame(), None, None, None

    header_row = all_rows[0]

    # Parse month column indices from header labels
    # Future:  'M1', 'M2', … → 1, 2, …
    # Historic: '−M1', '−M2', … (hyphen-minus or en-dash prefix) → −1, −2, …
    month_labels = [h for h in header_row[_COL_FIRST_M:] if h is not None]
    month_indices = _parse_month_labels(month_labels, historic=historic,
                                        sheet_name=sheet_name)

    if not month_indices:
        warn(
            f"Scenario sheet '{sheet_name}': no month columns found",
            source=f"scenario_loader/{sheet_name}",
        )
        return pd.DataFrame(), None, None, None

    n_months = len(month_indices)

    # Parse data rows
    run_id:   Any = None
    scn_name: Any = None
    no:       Any = None
    records: dict[str, list] = {}

    for row in all_rows[1:]:
        if row[_COL_ROW] is None:
            continue  # skip blank rows
        var_name = str(row[_COL_ROW]).strip()

        # Capture metadata from first data row
        if run_id is None:
            run_id   = row[_COL_RUN_ID]
            scn_name = row[_COL_SCN]
            no       = row[_COL_NO]

        # Extract monthly values (pad/truncate to match n_months)
        raw_values = list(row[_COL_FIRST_M:])
        values = (raw_values + [None] * n_months)[:n_months]
        records[var_name] = values

    if not records:
        warn(
            f"Scenario sheet '{sheet_name}': no data rows loaded",
            source=f"scenario_loader/{sheet_name}",
        )
        return pd.DataFrame(), run_id, scn_name, no

    df = pd.DataFrame(records, index=month_indices).T
    df.index.name = "variable"
    df.columns.name = "month"

    return df, run_id, scn_name, no


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_month_labels(
    labels: list,
    *,
    historic: bool,
    sheet_name: str,
) -> list[int]:
    """
    Convert ['M1','M2',…] → [1,2,…]  or  ['−M1','−M2',…] → [−1,−2,…].

    Tolerates both ASCII hyphen-minus (U+002D) and Unicode minus/en-dash
    prefixes that Excel sometimes introduces.
    """
    indices: list[int] = []
    for lbl in labels:
        s = str(lbl).strip()
        # Strip any leading minus / dash variants then strip 'M'
        sign = -1 if (s.startswith("-") or s.startswith("−") or s.startswith("–")) else 1
        digits = s.lstrip("-−–").lstrip("Mm")
        try:
            idx = sign * int(digits)
        except ValueError:
            warn(
                f"Scenario sheet '{sheet_name}': unrecognised month label {lbl!r} — skipped",
                source=f"scenario_loader/{sheet_name}",
            )
            continue
        indices.append(idx)
    return indices
