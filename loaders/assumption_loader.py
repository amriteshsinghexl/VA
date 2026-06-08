"""
§5 / §23.4 — Assumption Loader

Reads Assumptions_Extracted.xlsx and returns {sheet_name: DataFrame}.

Sheet structure (observed from workbook — 10-row metadata block):
  Row 1 : ['Assumption Table', '<name>']
  Row 2 : ['Source Range',     '<range>']
  Row 3 : ['Reserve Basis',    '<basis>']
  Row 4 : ['Assumption Type',  '<type>']
  Row 5 : ['Lookup Dimensions','<dims>']
  Row 6 : ['Consumed By',      '<consumer>']
  Row 7 : ['Source Anchors',   '<anchors>']
  Row 8 : blank
  Row 9 : ['DATA EXTRACTION (preserved verbatim …)', None, …]
  Row 10: blank
  Row 11: ['src row 4', <super-header or None>, …]   ← source row 4 content
  Row 12: ['src row 5', <col-name-0>, <col-name-1>, …] ← source row 5 — often column headers
  Row 13: ['src row 6', <val-0>, <val-1>, …]            ← source row 6 — first data row or headers
  Row 14: blank
  Row 15: ['src row', '<col_letter_1>', '<col_letter_2>', …]  ← source column letters (reference only)
  Row 16+: [<src_row_num>, <idx_val>, <data_1>, <data_2>, …] ← ACTUAL DATA

Column header resolution:
  • If row 12[1:] has ≥ 2 non-None values AND at least 1 string → use row 12[1:] as column headers.
  • Else if row 13[1:] has ≥ 2 non-None values → use row 13[1:] as column headers.
  • Else fall back to row 11[1:] or integer positional names.

Index:
  • Column 0 of the data block (src_row_num) is stored as 'src_row'.
  • Column 1 (first data value) is set as the DataFrame index when it is non-None across all rows.
  • DataFrame.attrs['meta'] carries the 7-key metadata dict.
  • Sheets with '[DUMMY placeholder RAND]' cells emit ONE ModelWarning per sheet; those
    cells become NaN.

D-002: '_README' is skipped; exactly 72 data DataFrames are returned.
"""
from __future__ import annotations

import math
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

from loaders.warnings import warn

_README = "_README"
# The workbook stores masked cells with a Unicode replacement char (U+FFFD) embedded:
# '[DUMMY � placeholder RAND]' — match on the invariant substrings instead.
_DUMMY_SUBSTR_A = "DUMMY"
_DUMMY_SUBSTR_B = "placeholder RAND"
_META_KEYS = [
    "assumption_table",
    "source_range",
    "reserve_basis",
    "assumption_type",
    "lookup_dims",
    "consumed_by",
    "source_anchors",
]

# Number of rows in the fixed metadata + header block before data begins
_DATA_START_ROW_IDX = 15   # row 16 (1-indexed) = index 15 (0-indexed)
_HDR_CTX_ROW_IDXS = (10, 11, 12)   # extracted rows 11, 12, 13 (0-indexed 10, 11, 12)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_assumptions(assumptions_path: str) -> dict[str, pd.DataFrame]:
    """
    Load all 72 data sheets from Assumptions_Extracted.xlsx.

    Returns {sheet_name: DataFrame} where each DataFrame has:
      • Column 'src_row' = source row number (for debugging / anchor lookup)
      • Remaining columns = data values with auto-resolved column headers
      • Index set from the first non-src_row column when values are uniform
      • .attrs['meta'] dict with the 7 metadata fields

    Skips '_README' (D-002).
    """
    try:
        wb = openpyxl.load_workbook(
            assumptions_path, data_only=True, read_only=True
        )
    except Exception as e:
        raise FileNotFoundError(
            f"Cannot open assumptions file '{assumptions_path}': {e}"
        ) from e

    result: dict[str, pd.DataFrame] = {}

    for name in wb.sheetnames:
        if name == _README:
            continue
        ws = wb[name]
        try:
            meta, df = _parse_sheet(ws, name)
        except Exception as exc:
            warn(
                f"Failed to parse assumption sheet '{name}': {exc} — sheet skipped",
                source=f"assumption_loader/{name}",
            )
            continue
        result[name] = df

    wb.close()

    if len(result) != 72:
        warn(
            f"Expected 72 data sheets in assumptions file, loaded {len(result)}",
            source="assumption_loader/sheet_count",
        )

    return result


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def _parse_sheet(
    ws, sheet_name: str
) -> tuple[dict[str, Any], pd.DataFrame]:
    """Parse one assumption sheet → (metadata_dict, DataFrame)."""

    all_rows: list[tuple] = list(
        ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
    )

    # ---- Metadata (rows 1-7, 0-indexed 0-6) --------------------------------
    meta: dict[str, Any] = {}
    for i, key in enumerate(_META_KEYS):
        row = all_rows[i] if i < len(all_rows) else ()
        meta[key] = row[1] if len(row) > 1 else None

    # ---- Header context rows (extracted rows 11-13, 0-indexed 10-12) -------
    hdr_ctx: list[tuple] = []
    for idx in _HDR_CTX_ROW_IDXS:
        row = all_rows[idx][1:] if idx < len(all_rows) else ()
        hdr_ctx.append(tuple(row))   # strip the 'src row X' label in col 0

    # Store src-row-4 content in meta for sheets that embed scalar anchors there
    # (e.g. mortality improvement base year, fee table reference year)
    row4_raw = hdr_ctx[0] if hdr_ctx else ()
    row4_values = [v for v in row4_raw if v is not None]
    meta['hdr_row4'] = row4_values[0] if row4_values else None

    # Resolve column headers
    col_headers = _resolve_col_headers(hdr_ctx, sheet_name)

    # ---- Data block (extracted rows 16+, 0-indexed 15+) -------------------
    has_dummy = False
    data_rows: list[list] = []

    for row in all_rows[_DATA_START_ROW_IDX:]:
        if not row or row[0] is None:
            continue  # skip trailing blank rows
        src_row_num = row[0]
        values = list(row[1:])   # drop src row number column

        # Detect and clean DUMMY placeholders (marker may contain U+FFFD replacement char)
        cleaned: list = []
        for v in values:
            if isinstance(v, str) and _DUMMY_SUBSTR_A in v and _DUMMY_SUBSTR_B in v:
                has_dummy = True
                cleaned.append(np.nan)
            else:
                cleaned.append(_coerce(v))
        data_rows.append([src_row_num] + cleaned)

    if has_dummy:
        warn(
            f"Sheet '{sheet_name}' contains masked placeholder cells "
            f"('[DUMMY … placeholder RAND]') — those cells are NaN. "
            f"Populate from live source before production use.",
            source=f"assumption_loader/{sheet_name}",
        )

    if not data_rows:
        df = pd.DataFrame()
        df.attrs["meta"] = meta
        return meta, df

    # ---- Build DataFrame ---------------------------------------------------
    n_data_cols = len(data_rows[0]) - 1   # excluding src_row col

    # Build column names: ['src_row'] + col_headers (padded/trimmed to n_data_cols)
    if col_headers and len(col_headers) >= n_data_cols:
        data_col_names = list(col_headers[:n_data_cols])
    elif col_headers:
        # Pad with positional indices for any extra columns beyond what headers cover
        data_col_names = list(col_headers) + list(range(len(col_headers), n_data_cols))
    else:
        data_col_names = list(range(n_data_cols))

    all_col_names = ["src_row"] + data_col_names
    df = pd.DataFrame(data_rows, columns=all_col_names)

    # ---- Set row index when first data column is a labelled string key --------
    # Rules:
    #   1. First column name must be a STRING (not an integer) — integer column
    #      names indicate dimension values (e.g. durations 1,2,3) not row keys.
    #   2. Column name must be unique in the DataFrame (no duplicate-name ambiguity).
    #   3. Index values must be all non-null AND unique (true row-key semantics).
    if n_data_cols > 0:
        first_data_col = data_col_names[0]
        is_string_key = isinstance(first_data_col, str)
        col_is_unique_name = data_col_names.count(first_data_col) == 1
        if is_string_key and col_is_unique_name:
            idx_series = df[first_data_col]
            try:
                if bool(idx_series.notna().all()) and idx_series.nunique() == len(df):
                    df = df.set_index(first_data_col)
            except (ValueError, TypeError):
                pass   # keep RangeIndex on any edge case

    df.attrs["meta"] = meta
    return meta, df


# ---------------------------------------------------------------------------
# Column header resolution
# ---------------------------------------------------------------------------

def _resolve_col_headers(
    hdr_ctx: list[tuple], sheet_name: str
) -> list:
    """
    Determine the best column header list from the three header context rows.

    hdr_ctx[0] = src row 4[1:]   (usually super-header or all-None)
    hdr_ctx[1] = src row 5[1:]   (often column headers)
    hdr_ctx[2] = src row 6[1:]   (sometimes column headers, sometimes 1st data row)

    Rules (applied in order, first match wins):
    1. Row 5 [hdr_ctx[1]]: ≥ 2 non-None values → use row 5 (integers ARE valid headers)
    2. Row 6 [hdr_ctx[2]]: ≥ 2 non-None values → use row 6
    3. Row 4 [hdr_ctx[0]]: ≥ 2 non-None values → use row 4
    4. Fall back to empty list (positional columns used by DataFrame constructor)

    Rationale: row 5 is ALWAYS the dimension-header row when it has values.
    The old string-only rule incorrectly skipped integer column headers
    (e.g. duration buckets 1,2,3,... for wide lapse-factor tables).
    """
    def non_none_count(row: tuple) -> int:
        return sum(1 for v in row if v is not None)

    # Rule 1: prefer row 5 unconditionally when it has ≥ 2 non-None values
    r5 = hdr_ctx[1] if len(hdr_ctx) > 1 else ()
    if non_none_count(r5) >= 2:
        return _strip_trailing_nones(list(r5))

    # Rule 2: row 6 with any 2+ non-None values
    r6 = hdr_ctx[2] if len(hdr_ctx) > 2 else ()
    if non_none_count(r6) >= 2:
        return _strip_trailing_nones(list(r6))

    # Rule 3: row 4 fallback
    r4 = hdr_ctx[0] if len(hdr_ctx) > 0 else ()
    if non_none_count(r4) >= 2:
        return _strip_trailing_nones(list(r4))

    return []


def _strip_trailing_nones(lst: list) -> list:
    while lst and lst[-1] is None:
        lst.pop()
    return lst


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _coerce(v: Any) -> Any:
    """Coerce values: float NaN → np.nan; leave everything else as-is."""
    if isinstance(v, float) and math.isnan(v):
        return np.nan
    return v
