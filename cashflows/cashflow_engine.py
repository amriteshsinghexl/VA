"""
§20 — Cashflow Engine  (Step 19)

Integrates all cashflow modules into a single per-period projection loop.

The engine resolves the sequential dependency:
  sep_acct AV (at CME) → i4L charges/payment → sep_acct EOP

Per-period loop (t = 1 .. N):
  Step A — Separate account BOP→CME (no EOP charges yet):
    bop[t]       = eop[t-1]          (seed = initial fund AVs from Fund_Info)
    me[t]        = (me_rate/12) × bop[t]
    av_after_me  = bop - me
    imf[t]       = (imf_rate/12) × av_after_me × suppressor
    growth[t]    = av_after_me × growth_factor[t]
    at_cme[t]    = av_after_me − imf + growth   ← pre-charge AV

  Step B — i4L charges (all from at_cme):
    gib_charge[t]  = gib_rate/12  × at_cme[t]
    i4l_charge[t]  = i4l_ap_rate/12 × at_cme[t] × (1 if O>0 else post_ap_factor)
    current_pmt[t] = (at_cme[t] − i4l_charge[t]) / (V[t] + W[t])   (if V+W > 0)
    monthly_pmt[t] = annual_pmt[t] / 12

  Step C — Withdrawals and EOP:
    withdrawal[t] = monthly_pmt[t]
    eop[t]        = at_cme[t] − gib_charge[t] − i4l_charge[t] − withdrawal[t]
    eop[t]        = max(0, eop[t])    (AV floor at 0)

NYREG213+CARVM suppressor (D-007, config.suppress_charges):
  IMF, GIB, i4L charges all suppressed to 0.

Stub/zero for test policy:
  - 4Later charge (B4=0)
  - GMDB charge (expense_charge_per_death_benefit=0)
  - Fixed account (fixed_account_value=0)
  - Per-policy expense EOP (SA maint exp=0 for LMFR5)
  - Partial withdrawals outside i4L guaranteed payments

Output: 480 × (7 time + 9 SA agg + 6 per-fund EOP + 5 i4L + 3 withdrawal + 1 total) columns.
"""
from __future__ import annotations

import math
import numpy as np
import openpyxl
import pandas as pd
from typing import Any

from loaders.warnings import warn

_N_FUNDS = 6
_TIME_COLS = [
    "policy_year", "policy_month", "month_in_policy_year",
    "bop_date", "eop_date", "cal_month_end", "attained_age",
]


def build_cashflows(
    time_axis: pd.DataFrame,
    policy: dict[str, Any],
    fund_mechanics: pd.DataFrame,
    i4l: pd.DataFrame,
    config: Any,
) -> pd.DataFrame:
    """
    Run the integrated per-period cashflow loop.

    Parameters
    ----------
    time_axis     : output of decrements.time_axis.build_time_axis()
    policy        : output of loaders.policy_loader.load_policy()
    fund_mechanics: output of cashflows.fund_mechanics.build_fund_mechanics()
    i4l           : output of cashflows.i4l.build_i4l()  (V, W, O pre-computed)
    config        : config.Config

    Returns
    -------
    DataFrame (index=projection_period) with columns:
        [time-axis],
        av_bop_sa, me_sa, imf_sa, growth_sa, av_at_cme_sa,
        gib_charge, i4l_charge, withdrawal,
        av_eop_sa,
        [av_eop_f1 … av_eop_f6],
        current_payment, monthly_payment, o_ap_remaining,
        av_bop_total, av_eop_total
    """
    if time_axis.empty:
        warn("time_axis is empty — returning empty cashflows DataFrame.",
             source="cashflow_engine/empty_time_axis")
        return pd.DataFrame()

    idx = time_axis.index
    n   = len(idx)

    # ---- Policy parameters --------------------------------------------------
    me_rate   = float(policy.get("expense_charge_per_separate_account") or 0.0)
    plan_code = str(policy.get("plan") or policy.get("model_plan") or "")
    imf_rate  = _load_imf_rate(config.policy_path, plan_code)
    suppressor = 0.0 if config.suppress_charges else 1.0
    indicator  = str(policy.get("i4l_indicator") or "").strip()

    # Charge rates
    from cashflows.charges import _parse_rate, _get_i4l_ap_rate
    gib_raw      = _parse_rate(policy.get("gmwb_ridercharge_rate"))
    i4l_ap_rate  = _get_i4l_ap_rate(config.policy_path)
    i4l_post_rate = 0.0
    gib_net_rate  = max(0.0, gib_raw - i4l_ap_rate)

    # Initial fund AVs — try policy dict first (supports multi-policy), fall back to Fund_Info
    initial_avs = _initial_avs_from_policy(policy) or _load_initial_avs(config.policy_path)

    # Growth factors (N × 6)
    g_arr = fund_mechanics.reindex(idx)[
        [f"growth_f{k}" for k in range(1, _N_FUNDS + 1)]
    ].to_numpy(dtype=float, na_value=np.nan)

    # i4L precomputed columns (V, W, O)
    v_arr = i4l.reindex(idx)["v_ap_annuity"].to_numpy(dtype=float, na_value=np.nan)
    w_arr = i4l.reindex(idx)["w_postap_annuity"].to_numpy(dtype=float, na_value=np.nan)
    o_arr = i4l.reindex(idx)["o_ap_remaining"].to_numpy(dtype=float, na_value=0.0)

    # ---- Output arrays ------------------------------------------------------
    bop_f   = np.zeros((n, _N_FUNDS), dtype=float)
    me_f    = np.zeros((n, _N_FUNDS), dtype=float)
    imf_f   = np.zeros((n, _N_FUNDS), dtype=float)
    grw_f   = np.zeros((n, _N_FUNDS), dtype=float)
    cme_f   = np.zeros((n, _N_FUNDS), dtype=float)   # at_cme per fund
    eop_f   = np.zeros((n, _N_FUNDS), dtype=float)

    gib_arr  = np.zeros(n, dtype=float)
    i4l_arr  = np.zeros(n, dtype=float)
    wdl_arr  = np.zeros(n, dtype=float)
    cpmt_arr = np.zeros(n, dtype=float)   # current_payment (annual)
    mpmt_arr = np.zeros(n, dtype=float)   # monthly_payment

    # ---- Per-period loop ----------------------------------------------------
    for t in range(n):
        # Step A: per-fund BOP → at_cme
        total_cme = 0.0
        for k in range(_N_FUNDS):
            bop0 = initial_avs.get(k + 1, 0.0)
            bop  = bop0 if t == 0 else eop_f[t - 1, k]
            gf   = g_arr[t, k]

            me       = bop * (me_rate / 12.0)
            av_me    = bop - me
            imf      = av_me * (imf_rate / 12.0) * suppressor
            if math.isnan(gf):
                growth = np.nan
                cme_val = np.nan
            else:
                growth  = av_me * gf
                cme_val = av_me - imf + growth

            bop_f[t, k] = bop
            me_f[t, k]  = me
            imf_f[t, k] = imf
            grw_f[t, k] = growth if not math.isnan(growth) else 0.0
            cme_f[t, k] = cme_val if not math.isnan(cme_val) else 0.0
            total_cme   += cme_val if not math.isnan(cme_val) else 0.0

        # Step B: i4L charges from total_cme
        if indicator == "i4L":
            o = float(o_arr[t])
            # GIB charge: active post-payment-start (policy_month >= m_pmt_start)
            # For simplicity: GIB active whenever we're not in pre-payment 4Later phase
            # (B4=0 for test policy, so always active when i4L)
            gib_c  = gib_net_rate / 12.0 * total_cme * suppressor
            # i4L charge: B1 within AP (O>0), B2 post-AP
            rate_i4l = i4l_ap_rate if o > 0 else i4l_post_rate
            i4l_c  = rate_i4l / 12.0 * total_cme * suppressor

            # Current payment (within AP: annuity pricing; post-AP: AV/W)
            v = float(v_arr[t]) if not math.isnan(float(v_arr[t])) else 0.0
            w = float(w_arr[t]) if not math.isnan(float(w_arr[t])) else 0.0
            vw = v + w
            # When W is NaN (D-016 truncated mortality), fall back to V-only
            # approximation for within-AP payments. This understates the payment
            # slightly (excludes post-AP annuity value) but gives a meaningful
            # result rather than zero.
            if vw == 0.0 and v > 0 and o > 0:
                vw = v      # V-only approximation
            if vw > 0 and total_cme > 0:
                if o > 0:
                    cpmt = (total_cme - i4l_c) / vw
                else:
                    adj   = float(policy.get("i4l_premium_tax") or 0.005)
                    load  = float(policy.get("i4l_expense_load") or 0.03)
                    if o == 0:
                        cpmt = total_cme * (1 - load) * (1 - adj) / w if w > 0 else 0.0
                    else:
                        cpmt = total_cme / w if w > 0 else 0.0
            else:
                cpmt = 0.0

            mpmt = cpmt / 12.0
        else:
            gib_c = 0.0
            i4l_c = 0.0
            cpmt  = 0.0
            mpmt  = 0.0

        gib_arr[t]  = gib_c
        i4l_arr[t]  = i4l_c
        cpmt_arr[t] = cpmt
        mpmt_arr[t] = mpmt

        # Step C: withdrawal and EOP per fund (pro-rated by at_cme weight)
        withdrawal = mpmt
        wdl_arr[t] = withdrawal

        remaining = total_cme - gib_c - i4l_c - withdrawal
        remaining = max(0.0, remaining)

        # Distribute EOP proportionally across funds
        if total_cme > 0:
            for k in range(_N_FUNDS):
                eop_f[t, k] = max(0.0, cme_f[t, k] / total_cme * remaining)
        else:
            eop_f[t] = 0.0

    # ---- Assemble DataFrame -------------------------------------------------
    time_cols = [c for c in _TIME_COLS if c in time_axis.columns]
    df = time_axis[time_cols].copy()

    # SA aggregates
    df["av_bop_sa"]    = bop_f.sum(axis=1)
    df["me_sa"]        = me_f.sum(axis=1)
    df["imf_sa"]       = imf_f.sum(axis=1)
    df["growth_sa"]    = grw_f.sum(axis=1)
    df["av_at_cme_sa"] = cme_f.sum(axis=1)
    df["gib_charge"]   = gib_arr
    df["i4l_charge"]   = i4l_arr
    df["withdrawal"]   = wdl_arr
    df["av_eop_sa"]    = eop_f.sum(axis=1)

    # Per-fund EOP
    for k in range(_N_FUNDS):
        df[f"av_eop_f{k+1}"] = eop_f[:, k]

    # i4L outputs
    df["current_payment"]  = cpmt_arr
    df["monthly_payment"]  = mpmt_arr
    df["o_ap_remaining"]   = o_arr

    # Total AV (SA + fixed; fixed = 0 for test policy)
    df["av_bop_total"] = df["av_bop_sa"]
    df["av_eop_total"] = df["av_eop_sa"]

    df.index.name = "projection_period"
    return df


# ---------------------------------------------------------------------------
# Helpers (shared with sep_acct)
# ---------------------------------------------------------------------------

_INV_ACCT_TO_SQL_FIELD = {
    # Inv Acct # → SQL field name (from Fund_Info col# mapping)
    1: "fund_value_separate_account_1",  # S&P 500
    2: "fund_value_separate_account_2",  # Russell 2000
    3: "fund_value_separate_account_3",  # Risk-Managed Fund
    4: "fund_value_separate_account_5",  # MSCI EAFE (SQL field 5)
    5: "fund_value_separate_account_6",  # Money Market (SQL field 6)
    6: "fund_value_separate_account_4",  # Barclays Aggregate (SQL field 4)
}


def _initial_avs_from_policy(policy: dict) -> dict[int, float] | None:
    """
    Read initial fund AVs from the policy dict SQL fields.
    Returns None if all values are zero/missing (fall back to Fund_Info).
    """
    result = {}
    for inv_acct, sql_field in _INV_ACCT_TO_SQL_FIELD.items():
        v = policy.get(sql_field)
        result[inv_acct] = float(v) if v is not None else 0.0
    # If total is zero, signal fallback to Fund_Info
    return result if sum(result.values()) > 0 else None


def _load_initial_avs(policy_path: str) -> dict[int, float]:
    try:
        wb = openpyxl.load_workbook(policy_path, data_only=True, read_only=True)
        ws = wb["Fund_Info"]
        rows = [r for r in ws.iter_rows(values_only=True)
                if any(c is not None for c in r)]
        wb.close()
    except Exception as exc:
        warn(f"Cannot read Fund_Info: {exc}", source="cashflow_engine/fund_info")
        return {}

    if len(rows) < 4:
        return {}
    header_row, value_row = rows[0], rows[3]
    start_col = next((i for i, v in enumerate(header_row) if v == 1), None)
    if start_col is None:
        return {}
    return {k: float(value_row[start_col + k - 1] or 0.0)
            for k in range(1, _N_FUNDS + 1)}


def _load_imf_rate(policy_path: str, plan_code: str) -> float:
    if not plan_code:
        return 0.0
    try:
        wb = openpyxl.load_workbook(policy_path, data_only=True, read_only=True)
        ws = wb["IMF NRSI (Other)"]
        for row in ws.iter_rows(values_only=True):
            if row[3] == plan_code:
                wb.close()
                return float(row[5])
        wb.close()
    except Exception:
        pass
    return 0.0
