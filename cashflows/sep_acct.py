"""
§12 — Separate Account AV Waterfall  (Step 13)

Projects the separate-account AV for each of the 6 investment accounts through
time using the roll-forward waterfall from Calc_SepAcct.

Simplified waterfall (correct for test policy 842612365):
  bop[t]        = eop[t-1]          (seed: initial fund AV from Fund_Info)
  me[t]         = (me_rate/12) × bop[t]
  av_after_me   = bop - me
  imf[t]        = (imf_rate/12) × av_after_me × carvm_suppressor
  growth[t]     = av_after_me × growth_factor[t]    (from fund_mechanics)
  av_at_cme[t]  = av_after_me − imf + growth
  ── stub charges (= 0 until their modules are built) ──
    partial_withdrawal, glwb_charge, 4later_charge,
    gib_charge, i4l_charge, i4l_policy_load, gmdb_i4l_charge
  eop[t]        = av_at_cme[t]      (because all stub charges = 0)

Output aggregate (sum of all 6 funds):
  av_bop_sa, me_sa, imf_sa, growth_sa, av_eop_sa

What is omitted (stub = 0):
  - Premiums (single-premium policy; all premium already in initial AV)
  - Rebalancing (allocations drift; all funds grow identically in test scenario)
  - Persistency bonus (persistency_credit_percent = 0 for test policy)
  - SA maintenance expense BOP (= 0 for plan LMFR5)
  - 401K charge (not a 401K plan)
  - DCA transfers (DCA balance = 0)
  - All i4L/GLWB/GIB/4Later charges (deferred to Steps 14-20)
  - GMDB charge non-i4L (= 0 because i4l_indicator = 'i4L')
  - Per-policy expense EOP (deferred)

CARVM suppressor (D-007):
  Applied to IMF charge.  When reserve_basis='NYREG213' and
  reserve_method='CARVM', suppress_charges=True → imf = 0.

D-005 (stub period = 0):
  Full monthly growth is applied BOP-to-CME; CME-to-EOP growth = 0.
  Because stub_period = 0 the CME coincides with EOP.

Data sources:
  - Initial fund AVs: Fund_Info sheet of the policy workbook (by Inv Acct #)
  - IMF rate:         IMF NRSI (Other) sheet (lookup by plan code)
  - M&E rate:         policy['expense_charge_per_separate_account'] (annual)
"""
from __future__ import annotations

import numpy as np
import openpyxl
import pandas as pd
from typing import Any

from loaders.warnings import warn

_N_FUNDS = 6

_FUND_NAMES = [
    "S&P 500",
    "Russell 2000",
    "Risk-Managed Fund",
    "MSCI EAFE",
    "Money Market",
    "Barclays Capital Aggregate",
]

_TIME_COLS = [
    "policy_year", "policy_month", "month_in_policy_year",
    "bop_date", "eop_date", "cal_month_end", "attained_age",
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_sep_acct(
    time_axis: pd.DataFrame,
    policy: dict[str, Any],
    fund_mechanics: pd.DataFrame,
    config: Any,
) -> pd.DataFrame:
    """
    Project separate-account AV through all projection periods.

    Parameters
    ----------
    time_axis      : output of decrements.time_axis.build_time_axis()
    policy         : output of loaders.policy_loader.load_policy()
    fund_mechanics : output of cashflows.fund_mechanics.build_fund_mechanics()
    config         : config.Config  (uses policy_path, suppress_charges,
                                     reserve_basis, reserve_method)

    Returns
    -------
    DataFrame (index=projection_period) with columns:
        [time-axis pass-through],
        av_bop_f1 … av_bop_f6,
        me_f1 … me_f6,
        imf_f1 … imf_f6,
        growth_f1 … growth_f6,
        av_eop_f1 … av_eop_f6,
        av_bop_sa, me_sa, imf_sa, growth_sa, av_eop_sa
    """
    if time_axis.empty:
        warn("time_axis is empty — returning empty sep_acct DataFrame.",
             source="sep_acct/empty_time_axis")
        return pd.DataFrame()

    # ---- parameters --------------------------------------------------------
    me_rate_annual = float(policy.get("expense_charge_per_separate_account", 0.0) or 0.0)
    me_rate_monthly = me_rate_annual / 12.0

    plan_code = str(policy.get("plan") or policy.get("model_plan") or "")
    imf_rate_annual = _load_imf_rate(config.policy_path, plan_code)
    imf_rate_monthly = imf_rate_annual / 12.0
    imf_suppressor = 0.0 if config.suppress_charges else 1.0

    from cashflows.cashflow_engine import _initial_avs_from_policy, _load_initial_avs as _ce_load
    initial_avs = _initial_avs_from_policy(policy) or _ce_load(config.policy_path)

    # ---- align fund_mechanics to time_axis index ---------------------------
    idx = time_axis.index
    n   = len(idx)

    growth_cols = [f"growth_f{k}" for k in range(1, _N_FUNDS + 1)]
    if not all(c in fund_mechanics.columns for c in growth_cols):
        warn(
            "fund_mechanics is missing expected growth_f1…f6 columns — "
            "returning NaN sep_acct DataFrame.",
            source="sep_acct/missing_growth_cols",
        )
        return pd.DataFrame()

    # growth factors array: shape (n, N_FUNDS)
    g_arr = fund_mechanics.reindex(idx)[growth_cols].to_numpy(dtype=float, na_value=np.nan)

    # ---- allocate output arrays (per fund) ---------------------------------
    bop_arr    = np.zeros((n, _N_FUNDS), dtype=float)
    me_arr     = np.zeros((n, _N_FUNDS), dtype=float)
    imf_arr    = np.zeros((n, _N_FUNDS), dtype=float)
    growth_arr = np.zeros((n, _N_FUNDS), dtype=float)
    eop_arr    = np.zeros((n, _N_FUNDS), dtype=float)

    for k in range(_N_FUNDS):
        inv_acct = k + 1          # Inv Acct # is 1-based
        bop0     = initial_avs.get(inv_acct, 0.0)

        for t in range(n):
            bop = bop0 if t == 0 else eop_arr[t - 1, k]
            gf  = g_arr[t, k]    # monthly growth factor

            me         = bop * me_rate_monthly
            av_after   = bop - me
            imf        = av_after * imf_rate_monthly * imf_suppressor

            if np.isnan(gf):
                growth  = np.nan
                eop_val = np.nan
            else:
                growth  = av_after * gf
                eop_val = av_after - imf + growth

            bop_arr[t, k]    = bop
            me_arr[t, k]     = me
            imf_arr[t, k]    = imf
            growth_arr[t, k] = growth
            eop_arr[t, k]    = eop_val

    # ---- assemble DataFrame ------------------------------------------------
    time_cols = [c for c in _TIME_COLS if c in time_axis.columns]
    df = time_axis[time_cols].copy()

    for k in range(_N_FUNDS):
        i = k + 1
        df[f"av_bop_f{i}"]  = bop_arr[:, k]
        df[f"me_f{i}"]      = me_arr[:, k]
        df[f"imf_f{i}"]     = imf_arr[:, k]
        df[f"growth_f{i}"]  = growth_arr[:, k]
        df[f"av_eop_f{i}"]  = eop_arr[:, k]

    # Aggregates
    df["av_bop_sa"]  = bop_arr.sum(axis=1)
    df["me_sa"]      = me_arr.sum(axis=1)
    df["imf_sa"]     = imf_arr.sum(axis=1)
    df["growth_sa"]  = growth_arr.sum(axis=1)
    df["av_eop_sa"]  = eop_arr.sum(axis=1)

    df.index.name = "projection_period"
    return df


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _load_initial_avs(policy_path: str) -> dict[int, float]:
    """
    Read initial fund AVs from Fund_Info sheet, keyed by Inv Acct # (1-6).

    Sheet layout:
      Row 0: (None, None, 'Inv Acct #', 1, 2, 3, 4, 5, 6, 7, 8, ...)
      Row 3: (None, None, 'Fund Value', av1, av2, av3, av4, av5, av6, ...)
    """
    try:
        wb = openpyxl.load_workbook(policy_path, data_only=True, read_only=True)
        ws = wb["Fund_Info"]
        rows = [r for r in ws.iter_rows(values_only=True)
                if any(c is not None for c in r)]
        wb.close()
    except Exception as exc:
        warn(f"Cannot read Fund_Info from '{policy_path}': {exc}",
             source="sep_acct/fund_info_error")
        return {}

    if len(rows) < 4:
        warn("Fund_Info sheet has fewer rows than expected — using zero initial AVs.",
             source="sep_acct/fund_info_short")
        return {}

    header_row = rows[0]   # ('Inv Acct #', 1, 2, 3, ...)
    value_row  = rows[3]   # ('Fund Value', av1, av2, ...)

    # Find the column where Inv Acct # 1 begins
    start_col = None
    for col_idx, val in enumerate(header_row):
        if val == 1:
            start_col = col_idx
            break

    if start_col is None:
        warn("Fund_Info: could not locate 'Inv Acct # 1' column.",
             source="sep_acct/fund_info_no_col1")
        return {}

    result: dict[int, float] = {}
    for inv_acct in range(1, _N_FUNDS + 1):
        col = start_col + (inv_acct - 1)
        try:
            result[inv_acct] = float(value_row[col] or 0.0)
        except (TypeError, ValueError, IndexError):
            result[inv_acct] = 0.0
            warn(f"Fund_Info: cannot read initial AV for Inv Acct {inv_acct}.",
                 source="sep_acct/fund_info_av_error")

    return result


def _load_imf_rate(policy_path: str, plan_code: str) -> float:
    """
    Read IMF annual rate from the 'IMF NRSI (Other)' sheet.

    Sheet layout (header row 0):
      TableKey | IAITMergeKey | Company | Plan | Mapping | Rate | NRSI Series

    Looks up by Plan == plan_code.  Returns 0.0 if not found.
    """
    if not plan_code:
        warn("No plan code provided — IMF rate defaults to 0.0.",
             source="sep_acct/imf_no_plan")
        return 0.0
    try:
        wb = openpyxl.load_workbook(policy_path, data_only=True, read_only=True)
        ws = wb["IMF NRSI (Other)"]
        for row in ws.iter_rows(values_only=True):
            if row[3] == plan_code:     # Col 3 = 'Plan'
                wb.close()
                return float(row[5])    # Col 5 = 'Rate'
        wb.close()
    except Exception as exc:
        warn(f"Cannot read IMF NRSI (Other) from '{policy_path}': {exc}",
             source="sep_acct/imf_error")

    warn(f"Plan '{plan_code}' not found in IMF NRSI (Other) — IMF rate = 0.0.",
         source="sep_acct/imf_plan_not_found")
    return 0.0
