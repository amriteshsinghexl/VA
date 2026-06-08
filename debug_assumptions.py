"""Debug script: find which sheets fail and why."""
import sys, traceback
sys.path.insert(0, ".")
import warnings; warnings.filterwarnings("ignore")
import openpyxl
import numpy as np

from loaders.assumption_loader import _parse_sheet

wb = openpyxl.load_workbook(
    r"data\Assumptions_Extracted.xlsx", data_only=True, read_only=True
)
failed = []
loaded = []
for name in wb.sheetnames:
    if name == "_README":
        continue
    ws = wb[name]
    try:
        meta, df = _parse_sheet(ws, name)
        loaded.append(name)
    except Exception as e:
        failed.append((name, str(e), traceback.format_exc()))

wb.close()

print(f"Loaded: {len(loaded)}  Failed: {len(failed)}")
print("\nFailed sheets:")
for name, err, tb in failed:
    print(f"\n  {name}: {err}")
    print(f"  {tb[-300:]}")

# Also check the 6 specific sheets we expect
expected_missing = [
    "VM21CA_BaseMortality_Single",
    "VM21CA_PM_BaseLapseRates",
    "NYREG213_StdScn_CARVM_InterestR",
]
for name in expected_missing:
    print(f"\n{name} in loaded: {name in loaded}")

# Check rows 11-16 of NYREG213_DynLapse_LifeFactors_M for column header debugging
wb2 = openpyxl.load_workbook(
    r"data\Assumptions_Extracted.xlsx", data_only=True, read_only=True
)
for sheet in ["NYREG213_DynLapse_LifeFactors_M", "VM21CA_BaseMortality_Single"]:
    if sheet not in wb2.sheetnames:
        print(f"\n{sheet}: NOT IN WORKBOOK")
        continue
    ws = wb2[sheet]
    print(f"\n=== {sheet} (rows={ws.max_row}, cols={ws.max_column}) ===")
    for i, row in enumerate(ws.iter_rows(min_row=11, max_row=17, values_only=True)):
        print(f"  Row {i+11}: {list(row[:8])}")
wb2.close()
