"""
Output Writer â€” Formatted Excel Workbook

Writes all model DataFrames to a single Excel workbook with professional
formatting: colour-coded tabs by layer, bold headers, auto-fit column widths,
frozen header rows, number/date/percentage format masks, and auto-filters.

Tab colour scheme
-----------------
  00_Policy_Summary          dark blue   (summary / inputs)
  01â€“04  Decrements layer    dark green
  05â€“09  Cashflows layer     dark orange
  10â€“14  Reserve layer       dark purple
  14_Reserve_Summary         dark red    (final output)
  Warnings                   grey
"""
from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from loaders.warnings import get_warnings


# ---------------------------------------------------------------------------
# Colour palette  (RGB hex, no leading #)
# ---------------------------------------------------------------------------

class _C:
    # Tab colours
    TAB_SUMMARY  = "1F4E79"   # dark navy blue
    TAB_DECR     = "375623"   # dark green
    TAB_CF       = "833C00"   # dark burnt orange
    TAB_RESERVE  = "4B1C8C"   # dark purple
    TAB_FINAL    = "C00000"   # dark red
    TAB_WARN     = "7F7F7F"   # grey

    # Header fills
    HDR_SUMMARY  = "BDD7EE"   # light blue
    HDR_DECR     = "E2EFDA"   # light green
    HDR_CF       = "FCE4D6"   # light orange
    HDR_RESERVE  = "E8DAEF"   # light lavender
    HDR_FINAL    = "FFCCCC"   # light red/pink
    HDR_WARN     = "F2F2F2"   # light grey

    # Header font colour (dark text on light background)
    HDR_FONT     = "000000"   # black

    # Alternating row fill (very light, every other data row)
    ROW_ALT      = "F7F7F7"   # near-white grey


# ---------------------------------------------------------------------------
# Sheet catalogue: (sheet_name, tab_colour, header_fill)
# ---------------------------------------------------------------------------

_SHEET_CATALOGUE = {
    "00_Policy_Summary": (_C.TAB_SUMMARY,  _C.HDR_SUMMARY),
    "01_Time_Axis":      (_C.TAB_DECR,     _C.HDR_DECR),
    "02_Mortality":      (_C.TAB_DECR,     _C.HDR_DECR),
    "03_Lapse":          (_C.TAB_DECR,     _C.HDR_DECR),
    "04_Lives":          (_C.TAB_DECR,     _C.HDR_DECR),
    "05_Interest_Rates": (_C.TAB_CF,       _C.HDR_CF),
    "06_Fund_Mechanics": (_C.TAB_CF,       _C.HDR_CF),
    "07_Sep_Acct":       (_C.TAB_CF,       _C.HDR_CF),
    "08_i4L_Rider":      (_C.TAB_CF,       _C.HDR_CF),
    "09_Cashflow_Engine":(_C.TAB_CF,       _C.HDR_CF),
    "10_Dec_Cashflows":  (_C.TAB_RESERVE,  _C.HDR_RESERVE),
    "11_StdScn_ANR":     (_C.TAB_RESERVE,  _C.HDR_RESERVE),
    "12_CARVM":          (_C.TAB_RESERVE,  _C.HDR_RESERVE),
    "13_DAC":            (_C.TAB_RESERVE,  _C.HDR_RESERVE),
    "14_Reserve_Summary":(_C.TAB_FINAL,    _C.HDR_FINAL),
    "Warnings":          (_C.TAB_WARN,     _C.HDR_WARN),
}


# ---------------------------------------------------------------------------
# Number format rules  (applied by column-name pattern)
# ---------------------------------------------------------------------------

_FMT_CURRENCY   = '#,##0.00'
_FMT_RATE6      = '0.000000'
_FMT_RATE4      = '0.0000'
_FMT_PCT        = '0.00%'
_FMT_INTEGER    = '0'
_FMT_DATE       = 'YYYY-MM-DD'

# Patterns matched against lowercase column name (first match wins)
_COL_FORMATS = [
    # Dates â€” match only actual date columns (not 'av_at_cme_sa' etc.)
    (r'_date$|^cal_month',              _FMT_DATE),
    # Integer-like
    (r'period|policy_yr|policy_month|month_in|attained|year|flag|'
     r'ap_remaining|ap_end|pmt_start|ap_date',
                                        _FMT_INTEGER),
    # Currency / dollar amounts
    (r'av_|fund|charge|payment|withdrawal|reserve|csv|pv_|cme|'
     r'anr|dac|benefit|base|income|accum',
                                        _FMT_CURRENCY),
    # Small rates / factors (6 dp)
    (r'q_|rate|factor|monthly|growth|disc|surv|mort|lapse|'
     r'suppressor|i4l_ap|i4l_post|gib|imf|me_|bey|aey',
                                        _FMT_RATE6),
    # Percentages stored as decimals (already 0â€“1 range)
    (r'pct|percent',                    _FMT_PCT),
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_output(
    *,
    policy: dict[str, Any],
    config: Any,
    time_axis: pd.DataFrame,
    mortality: pd.DataFrame,
    lapse: pd.DataFrame,
    lives: pd.DataFrame,
    interest_rates: pd.DataFrame,
    fund_mechanics: pd.DataFrame,
    sep_acct: pd.DataFrame,
    i4l: pd.DataFrame,
    cashflows: pd.DataFrame,
    dec_cf: pd.DataFrame,
    std_scn: pd.DataFrame,
    carvm: pd.DataFrame,
    dac: pd.DataFrame,
    reserve: pd.DataFrame,
    output_dir: str,
    filename: str = "abc_corp_va_output.xlsx",
) -> Path:
    """
    Write all model DataFrames to a formatted Excel workbook.

    Returns the Path to the written file.
    """
    out_dir  = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename

    wb = Workbook()
    wb.remove(wb.active)          # remove default blank sheet

    # ---- Summary sheet -------------------------------------------------------
    _write_summary_sheet(wb, policy, config, reserve)

    # ---- Projection sheets ---------------------------------------------------
    sheets = [
        ("01_Time_Axis",       time_axis),
        ("02_Mortality",       mortality),
        ("03_Lapse",           lapse),
        ("04_Lives",           lives),
        ("05_Interest_Rates",  interest_rates),
        ("06_Fund_Mechanics",  fund_mechanics),
        ("07_Sep_Acct",        sep_acct),
        ("08_i4L_Rider",       i4l),
        ("09_Cashflow_Engine", cashflows),
        ("10_Dec_Cashflows",   dec_cf),
        ("11_StdScn_ANR",      std_scn),
        ("12_CARVM",           carvm),
        ("13_DAC",             dac),
        ("14_Reserve_Summary", reserve),
    ]
    for name, df in sheets:
        _write_data_sheet(wb, name, df)

    # ---- Warnings sheet ------------------------------------------------------
    _write_warnings_sheet(wb)

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_summary_sheet(
    wb: Workbook,
    policy: dict[str, Any],
    config: Any,
    reserve: pd.DataFrame,
) -> None:
    """Policy Summary â€” two-column key/value layout with section dividers."""
    tab_colour, hdr_fill = _SHEET_CATALOGUE["00_Policy_Summary"]
    ws = wb.create_sheet("00_Policy_Summary")
    ws.sheet_properties.tabColor = tab_colour

    reserve_t0 = 0.0
    anr_t0     = 0.0
    if reserve is not None and not reserve.empty:
        if "reserve_t0" in reserve.columns:
            reserve_t0 = float(reserve["reserve_t0"].iloc[0])
        if "anr" in reserve.columns:
            anr_t0 = float(reserve["anr"].iloc[0])

    def _section(label: str):
        row = [label, ""]
        ws.append(row)
        r = ws.max_row
        for col in (1, 2):
            cell = ws.cell(r, col)
            cell.fill    = PatternFill("solid", fgColor=hdr_fill)
            cell.font    = Font(bold=True, color=_C.HDR_FONT, size=11)
            cell.alignment = Alignment(horizontal="left")

    def _row(key: str, value: Any, fmt: str | None = None):
        ws.append([key, value])
        r = ws.max_row
        ws.cell(r, 1).font      = Font(bold=True, size=10)
        ws.cell(r, 1).alignment = Alignment(horizontal="left")
        ws.cell(r, 2).alignment = Alignment(horizontal="left")
        if fmt:
            ws.cell(r, 2).number_format = fmt

    def _blank():
        ws.append(["", ""])

    # Title
    ws.append(["Abc_corp VA Python Valuation Model", ""])
    r = ws.max_row
    ws.cell(r, 1).font      = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(r, 1).fill      = PatternFill("solid", fgColor=tab_colour)
    ws.cell(r, 1).alignment = Alignment(horizontal="left")
    ws.cell(r, 2).fill      = PatternFill("solid", fgColor=tab_colour)

    _blank()

    _section("Policy")
    _row("Policy number",     policy.get("policy_number", ""))
    _row("Plan",              policy.get("plan") or policy.get("model_plan", ""))
    _row("Issue date",        str(policy.get("issue_date", "")))
    _row("Issue age",         policy.get("issue_age", ""))
    _row("Gender",            policy.get("gender1") or policy.get("gender", ""))
    _row("i4L indicator",     policy.get("i4l_indicator", ""))
    _row("DB option",         policy.get("deathbenefittype", ""))
    _blank()

    _section("Values at Valuation Date")
    _row("Total account value",   policy.get("total_account_value", 0.0),  _FMT_CURRENCY)
    _row("Cash surrender value",  policy.get("cash_surrender_value", 0.0), _FMT_CURRENCY)
    _row("i4L income base",       policy.get("4later_current_income_base", 0.0), _FMT_CURRENCY)
    _blank()

    _section("Run Configuration")
    _row("Reserve basis",    config.reserve_basis)
    _row("Reserve method",   config.reserve_method)
    _row("Projection months",config.projection_months)
    _row("Run date",         datetime.date.today().isoformat())
    _blank()

    _section("Results")
    _row("FINAL RESERVE (t=0)",  reserve_t0, _FMT_CURRENCY)
    _row("StdScn ANR (t=0)",     anr_t0,     _FMT_CURRENCY)
    _blank()

    _section("Output Sheets")
    sheet_guide = [
        ("00_Policy_Summary",   "This sheet â€” key fields and final reserve"),
        ("01_Time_Axis",        "Projection time spine (480 periods)"),
        ("02_Mortality",        "Monthly q, improvement multiplier, q_monthly"),
        ("03_Lapse",            "ITM bucket, SC flag, q_lapse_annual/monthly"),
        ("04_Lives",            "lives_bop, lives_eop (decrement spine)"),
        ("05_Interest_Rates",   "BEY, AEY, disc_factor (unshocked + shocked)"),
        ("06_Fund_Mechanics",   "Monthly growth factors per fund (f1â€“f6)"),
        ("07_Sep_Acct",         "Sep account AV waterfall â€” per fund + aggregate"),
        ("08_i4L_Rider",        "AP timing, discount U, annuity V/W, charges"),
        ("09_Cashflow_Engine",  "Integrated AV loop: at_cme, charges, EOP"),
        ("10_Dec_Cashflows",    "Lives-weighted cashflows (dec_cf = cf Ã— lives)"),
        ("11_StdScn_ANR",       "Accumulated fund/charges, net revenue, ANR"),
        ("12_CARVM",            "PV(CSV), PV(annuity), CARVM reserve per period"),
        ("13_DAC",              "Lives progression + DAC balance"),
        ("14_Reserve_Summary",  "Final binding reserve per period"),
        ("Warnings",            "All ModelWarnings raised during the run"),
    ]
    for sheet, desc in sheet_guide:
        ws.append([sheet, desc])
        ws.cell(ws.max_row, 1).font      = Font(bold=False, size=9, color="1F4E79")
        ws.cell(ws.max_row, 2).font      = Font(size=9)
        ws.cell(ws.max_row, 1).alignment = Alignment(horizontal="left")
        ws.cell(ws.max_row, 2).alignment = Alignment(horizontal="left")

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # Thin outer border on the results block
    _apply_thin_border(ws, row_start=1, row_end=ws.max_row, col_start=1, col_end=2)


def _write_data_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    """Write one projection DataFrame with full formatting."""
    tab_colour, hdr_fill = _SHEET_CATALOGUE.get(name, (_C.TAB_WARN, _C.HDR_WARN))
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_colour

    if df is None or df.empty:
        ws.append(["(no data)"])
        return

    # Reset index so projection_period becomes column 1
    df_out = df.reset_index()
    columns = list(df_out.columns)

    # ---- Header row ---------------------------------------------------------
    hdr_fill_obj  = PatternFill("solid", fgColor=hdr_fill)
    hdr_font_obj  = Font(bold=True, color=_C.HDR_FONT, size=10)
    hdr_align_obj = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_bottom   = Border(bottom=Side(style="medium", color="000000"))

    ws.append(columns)
    ws.row_dimensions[1].height = 30
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(1, col_idx)
        cell.fill      = hdr_fill_obj
        cell.font      = hdr_font_obj
        cell.alignment = hdr_align_obj
        cell.border    = thin_bottom

    # ---- Data rows ----------------------------------------------------------
    alt_fill = PatternFill("solid", fgColor=_C.ROW_ALT)

    for row_idx, row_data in enumerate(df_out.itertuples(index=False), 2):
        ws.append(list(row_data))
        is_alt = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row_idx, col_idx)
            # Alternating row shading
            if is_alt:
                cell.fill = alt_fill
            # Number format
            fmt = _col_format(col_name)
            if fmt:
                cell.number_format = fmt
            # Alignment
            cell.alignment = Alignment(horizontal="right" if _is_numeric_col(df_out, col_name)
                                       else "left", vertical="center")

    # ---- Column widths (auto-fit based on content) --------------------------
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        # Sample max length: header + up to 50 data rows
        sample_vals = df_out[col_name].head(50).astype(str)
        max_len = max(
            len(str(col_name)),
            sample_vals.str.len().max() if not sample_vals.empty else 0,
        )
        # Clamp: min 8, max 30 chars; wider for date/description columns
        if "date" in col_name.lower() or "cme" in col_name.lower():
            width = 14
        else:
            width = min(max(max_len + 2, 8), 30)
        ws.column_dimensions[col_letter].width = width

    # ---- Freeze header + projection_period column ---------------------------
    ws.freeze_panes = "B2"

    # ---- Auto-filter on header row ------------------------------------------
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _write_warnings_sheet(wb: Workbook) -> None:
    """Warnings sheet."""
    tab_colour, hdr_fill = _SHEET_CATALOGUE["Warnings"]
    ws = wb.create_sheet("Warnings")
    ws.sheet_properties.tabColor = tab_colour

    warns = get_warnings()
    cols  = ["source", "message"]

    hdr_fill_obj = PatternFill("solid", fgColor=hdr_fill)
    hdr_font_obj = Font(bold=True, color=_C.HDR_FONT, size=10)

    ws.append(cols)
    ws.row_dimensions[1].height = 22
    for col_idx in range(1, 3):
        cell = ws.cell(1, col_idx)
        cell.fill      = hdr_fill_obj
        cell.font      = hdr_font_obj
        cell.alignment = Alignment(horizontal="center")

    if not warns:
        ws.append(["(none)", "No warnings raised."])
    else:
        for w in warns:
            ws.append([w.source, w.message])
            ws.cell(ws.max_row, 2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:B1"

    # Row heights for wrapped messages
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 28


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _col_format(col_name: str) -> str | None:
    """Return an Excel number-format string for a column name, or None."""
    lower = col_name.lower()
    for pattern, fmt in _COL_FORMATS:
        if re.search(pattern, lower):
            return fmt
    return None


def _is_numeric_col(df: pd.DataFrame, col_name: str) -> bool:
    """True if the column holds numeric (non-date, non-string) data."""
    if col_name not in df.columns:
        return False
    dtype = df[col_name].dtype
    return pd.api.types.is_numeric_dtype(dtype)


def _apply_thin_border(
    ws, row_start: int, row_end: int, col_start: int, col_end: int
) -> None:
    """Apply a thin border around all cells in the given range."""
    thin = Side(style="thin", color="CCCCCC")
    for row in ws.iter_rows(min_row=row_start, max_row=row_end,
                             min_col=col_start, max_col=col_end):
        for cell in row:
            cell.border = Border(
                top    = thin if cell.row == row_start else Side(),
                bottom = thin if cell.row == row_end   else Side(),
                left   = thin if cell.column == col_start else Side(),
                right  = thin if cell.column == col_end  else Side(),
            )
