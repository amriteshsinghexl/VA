"""
Standalone FastAPI service for VA assumption tables.

Reads Assumptions_Extracted.xlsx (73 sheets: _README + 72 data sheets).
Each data sheet contains a 7-row metadata header followed by table data.

Usage:
    python -m uvicorn assumptions_api:app --port 8013 --reload
    -- or via run_all.py --

Endpoints:
    GET  /                       API root
    GET  /tables                 List all 72 assumption table sheet names
    GET  /tables/{name}          Full sheet data as JSON (paginated)
    GET  /tables/{name}/meta     Metadata header (rows 1â€“7) of a sheet
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from fastapi import FastAPI, HTTPException, Query

ASSUMPTIONS_FILE = (
    Path(__file__).resolve().parent.parent / "abc_corp_va" / "data" / "Assumptions_Extracted.xlsx"
)

app = FastAPI(
    title="VA Assumptions API",
    description="REST API for the Abc_corp VA assumption tables workbook",
    version="1.0.0",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sheet_names() -> List[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ASSUMPTIONS_FILE), read_only=True, data_only=True)
        names = [n for n in wb.sheetnames if not n.startswith("_")]
        wb.close()
        return names
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not open assumptions file: {exc}")


def _read_sheet(name: str) -> List[List[Any]]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ASSUMPTIONS_FILE), read_only=True, data_only=True)
        if name not in wb.sheetnames:
            available = _sheet_names()
            wb.close()
            raise HTTPException(
                status_code=404,
                detail=f"Sheet '{name}' not found. Available: {available}",
            )
        ws = wb[name]
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read sheet '{name}': {exc}")


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.get("/")
def root():
    exists = ASSUMPTIONS_FILE.exists()
    tables = _sheet_names() if exists else []
    return {
        "message": "VA Assumptions API",
        "assumptions_file": str(ASSUMPTIONS_FILE),
        "file_exists": exists,
        "table_count": len(tables),
        "tables": tables,
    }


@app.get("/tables")
def list_tables():
    """List the names of all 72 assumption table sheets."""
    tables = _sheet_names()
    return {"table_count": len(tables), "tables": tables}


@app.get("/tables/{name}")
def get_table(
    name: str,
    skip: int = Query(0, ge=0, description="Data rows to skip (after metadata header)"),
    limit: int = Query(100, ge=1, le=10000, description="Max data rows to return"),
):
    """Return paginated data rows from an assumption sheet.

    Rows 1â€“7 are the metadata header; data begins at row 8.
    """
    rows = _read_sheet(name)
    # Rows 0â€“6 are metadata (7 rows); data starts at row index 7
    META_ROWS = 7
    data_rows = rows[META_ROWS:]
    if not data_rows:
        return {"table": name, "total": 0, "skip": skip, "limit": limit, "headers": [], "data": []}

    # Row 8 (index 7 after slicing = index 0 here) contains column headers
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data_rows[0])]
    body = data_rows[1:]
    total = len(body)
    page = body[skip: skip + limit]
    data = [dict(zip(headers, [str(v) if v is not None else None for v in r])) for r in page]
    return {"table": name, "total": total, "skip": skip, "limit": limit, "headers": headers, "data": data}


@app.get("/tables/{name}/meta")
def get_table_meta(name: str):
    """Return the 7-row metadata header of an assumption sheet."""
    rows = _read_sheet(name)
    meta = [[str(c) if c is not None else None for c in row] for row in rows[:7]]
    return {"table": name, "metadata_rows": meta}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("assumptions_api:app", host="0.0.0.0", port=8013, reload=True)
