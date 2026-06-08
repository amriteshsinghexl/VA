"""
Standalone FastAPI service for VA model output files.

Serves the Excel workbooks written to abc_corp_va/results/ and exposes
sheet-level data for browsing reserve and cashflow outputs.

Usage:
    python -m uvicorn outputs_api:app --port 8014 --reload
    -- or via run_all.py --

Endpoints:
    GET  /                          API root
    GET  /files                     List all Excel result files
    GET  /files/{filename}          File metadata + sheet list
    GET  /files/{filename}/sheet/{sheet}   Sheet data as JSON (paginated)
    GET  /download/{filename}       Download the Excel workbook
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse

RESULTS_DIR = Path(__file__).resolve().parent.parent / "abc_corp_va" / "results"

app = FastAPI(
    title="VA Outputs API",
    description="REST API for Abc_corp VA model result workbooks",
    version="1.0.0",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe(filename: str) -> Path:
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    p = RESULTS_DIR / filename
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"File '{filename}' not found")
    return p


def _xlsx_sheets(path: Path) -> list[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not open workbook: {exc}")


def _read_sheet(path: Path, sheet: str, skip: int, limit: int) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            available = wb.sheetnames
            wb.close()
            raise HTTPException(
                status_code=404,
                detail=f"Sheet '{sheet}' not found. Available: {available}",
            )
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read sheet: {exc}")

    if not rows:
        return {"total": 0, "skip": skip, "limit": limit, "headers": [], "data": []}

    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    total = len(data_rows)
    page = data_rows[skip: skip + limit]
    data = [dict(zip(headers, [str(v) if v is not None else None for v in r])) for r in page]
    return {"total": total, "skip": skip, "limit": limit, "headers": headers, "data": data}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.get("/")
def root():
    files = [f.name for f in sorted(RESULTS_DIR.glob("*.xlsx"))] if RESULTS_DIR.exists() else []
    return {
        "message": "VA Outputs API",
        "results_dir": str(RESULTS_DIR),
        "files": files,
    }


@app.get("/files")
def list_files():
    """List all Excel result workbooks."""
    if not RESULTS_DIR.exists():
        return {"results_dir": str(RESULTS_DIR), "files": []}
    files = [
        {
            "filename": f.name,
            "size_bytes": f.stat().st_size,
            "download_url": f"/download/{f.name}",
        }
        for f in sorted(RESULTS_DIR.glob("*.xlsx"))
        if f.is_file()
    ]
    return {"results_dir": str(RESULTS_DIR), "files": files}


@app.get("/files/{filename}")
def describe_file(filename: str):
    """Return file metadata and list of worksheet names."""
    p = _safe(filename)
    sheets = _xlsx_sheets(p)
    return {
        "filename": filename,
        "size_bytes": p.stat().st_size,
        "sheets": sheets,
        "sheet_count": len(sheets),
    }


@app.get("/files/{filename}/sheet/{sheet}")
def get_sheet_data(
    filename: str,
    sheet: str,
    skip: int = Query(0, ge=0, description="Rows to skip (after header)"),
    limit: int = Query(100, ge=1, le=5000, description="Max rows to return"),
):
    """Return paginated rows from a named worksheet."""
    p = _safe(filename)
    return _read_sheet(p, sheet, skip, limit)


@app.get("/download/{filename}")
def download_file(filename: str):
    """Download a result Excel workbook."""
    p = _safe(filename)
    return FileResponse(
        path=str(p),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("outputs_api:app", host="0.0.0.0", port=8014, reload=True)
